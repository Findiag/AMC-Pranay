"""
claude_pdf_extractor.py — Claude-based PDF -> _extracted.xlsx fallback.

When the rule-based extractor in `extract_tables.py` produces structurally
broken output (CY column empty, watermark text in data, OCR garbling), this
module sends the original PDF straight to Claude with native PDF input and
extracts BS/PL tables directly. Output mirrors the shape produced by
`extract_tables.write_sheet` so the downstream mapper is unchanged.

Public API:
    is_extraction_broken(xlsx_path) -> tuple[bool, str]
        Returns (broken, reason). Heuristics-based.
    extract_with_claude(pdf_path, output_dir, api_key=None, model=None)
                                                            -> str | None
        Calls Claude with the PDF, writes _extracted.xlsx, returns its path.
        Returns None on failure (caller keeps the original extracted xlsx).

Both functions are import-safe: missing `anthropic` SDK or no API key just
makes `extract_with_claude` return None — the caller falls back to the
existing extracted xlsx silently.
"""
from __future__ import annotations

import base64
import json
import os
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import openpyxl  # required everywhere — already in requirements.txt
except ImportError:  # pragma: no cover
    openpyxl = None  # type: ignore

# Reuse the rule-based writer so the Claude-produced xlsx is byte-identical
# in shape to what the rest of the pipeline already consumes.
from extract_tables import (
    write_sheet, STMT_NAMES, normalize_rows_to_lakhs,
)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Heuristics: was the rule-based extraction broken enough to retry?
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Single-token design vocabulary that should never appear as a value
_DESIGN_TOKENS = {
    'creating', 'value', 'creation', 'accelerating', 'our', 'growth',
    'sustainability', 'statutory', 'reports', 'report', 'financial',
    'statements', 'annual', 'corporate', 'overview', 'governance',
}


def _looks_numeric(val) -> bool:
    if isinstance(val, (int, float)):
        return True
    if not isinstance(val, str):
        return False
    s = val.strip().replace(',', '').replace('(', '').replace(')', '').replace(' ', '')
    return bool(re.match(r'^-?\d+(\.\d+)?$', s))


def _to_float_or_none(v):
    if isinstance(v, (int, float)):
        return float(v)
    if not isinstance(v, str):
        return None
    s = v.strip().replace(',', '').replace(' ', '')
    if not s or s in ('-', '–', '—'):
        return None
    m = re.match(r'^\(([\d.]+)\)$', s)
    if m:
        try: return -float(m.group(1))
        except ValueError: return None
    try:
        return float(s)
    except ValueError:
        return None


def _ocr_garbled(text: str) -> bool:
    """Heuristic: a label like 'IInnccoommee' has 3+ runs of doubled
    consonant-pairs, which never happens in real English."""
    if not isinstance(text, str) or len(text) < 6:
        return False
    return len(re.findall(r'([A-Za-z])\1', text)) >= 4


def is_extraction_broken(xlsx_path: str) -> Tuple[bool, str]:
    """Inspect the rule-based extracted xlsx and decide whether the Claude
    fallback should run. The bias is INTENTIONALLY toward false positives —
    a needless Claude call costs a few rupees, but a silently-wrong report
    propagates to Airtable and downstream observations.

    Triggers (any one):
      1. CY column mostly empty (>70% blank among labelled data rows).
      2. PY column mostly empty when CY isn't (one-sided extraction).
      3. Note column contains numeric values with commas (CY leaked into Note).
      4. Any cell contains a single design-vocabulary watermark token.
      5. CY median magnitude differs from PY median by >50× (one column got
         note numbers instead of values).
      6. ≥3 data rows have OCR character-doubling in their label
         ("IInnccoommee"-style — Pritish_Nandi, Sicagen).
      7. >40% of CY values are suspiciously large (>10⁹), suggesting
         column-merge corruption ("4444248399" = "4,444" + "248,399" merged).
      8. No sheet has a labelled "Particulars" column at all (extractor
         failed to find any tables).
      9. The xlsx has no sheets / fails to open.

    Returns (broken, reason).
    """
    if openpyxl is None:
        return False, "openpyxl unavailable; skipping check"

    path = Path(xlsx_path)
    if not path.exists():
        return True, f"extracted xlsx missing: {path}"

    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
    except Exception as e:
        return True, f"failed to open extracted xlsx: {e}"

    if not wb.sheetnames:
        return True, "extracted xlsx has no sheets"

    sheets_with_data = 0

    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.max_row < 5 or ws.max_column < 4:
            continue

        header_row = None
        for r in range(1, min(ws.max_row, 10) + 1):
            v = ws.cell(row=r, column=1).value
            if isinstance(v, str) and v.strip().lower() == 'particulars':
                header_row = r
                break
        if header_row is None:
            continue

        sheets_with_data += 1
        data_rows = list(range(header_row + 1, ws.max_row + 1))
        if len(data_rows) < 5:
            continue

        # (4) Watermark token in any cell
        for r in data_rows:
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str):
                    s = v.strip().lower()
                    if s and len(s.split()) <= 2 and all(t in _DESIGN_TOKENS
                                                          for t in s.split()):
                        return True, (f"watermark token {s!r} in {sn} "
                                      f"R{r}C{c}")

        cy_present = 0
        py_present = 0
        cy_total = 0
        note_numeric = 0
        ocr_hits = 0
        cy_magnitudes: List[float] = []
        py_magnitudes: List[float] = []
        oversize_cy = 0

        for r in data_rows:
            part = ws.cell(row=r, column=1).value
            if not (isinstance(part, str) and part.strip()):
                continue
            cy_total += 1

            if _ocr_garbled(part):
                ocr_hits += 1

            cy_val_raw = ws.cell(row=r, column=3).value
            py_val_raw = ws.cell(row=r, column=4).value
            note_val = ws.cell(row=r, column=2).value

            cy_f = _to_float_or_none(cy_val_raw)
            py_f = _to_float_or_none(py_val_raw)
            if cy_f is not None:
                cy_present += 1
                cy_magnitudes.append(abs(cy_f))
                if abs(cy_f) > 1e9:
                    oversize_cy += 1
            if py_f is not None:
                py_present += 1
                py_magnitudes.append(abs(py_f))

            if isinstance(note_val, str) and ',' in note_val and \
                    _looks_numeric(note_val):
                note_numeric += 1

        if cy_total < 8:
            continue

        # (1)
        if cy_present / cy_total < 0.30:
            return True, (f"{sn}: only {cy_present}/{cy_total} data rows have "
                          f"a CY value (<30%) — column-detection failure")
        # (2)
        if py_present / cy_total < 0.30 and cy_present / cy_total > 0.60:
            return True, (f"{sn}: PY column nearly empty ({py_present}/"
                          f"{cy_total}) while CY is populated — one-sided "
                          f"extraction")
        # (3)
        if note_numeric >= 5:
            return True, (f"{sn}: {note_numeric} note cells contain comma-"
                          f"formatted numerics — CY values leaked into Note")
        # (5)
        if cy_magnitudes and py_magnitudes:
            cy_med = sorted(cy_magnitudes)[len(cy_magnitudes)//2]
            py_med = sorted(py_magnitudes)[len(py_magnitudes)//2]
            if cy_med > 0 and py_med > 0:
                ratio = max(cy_med, py_med) / min(cy_med, py_med)
                if ratio > 50:
                    return True, (f"{sn}: CY median ({cy_med:.0f}) and PY "
                                   f"median ({py_med:.0f}) differ by "
                                   f"{ratio:.0f}× — likely column-shift")
        # (6)
        if ocr_hits >= 3:
            return True, (f"{sn}: {ocr_hits} labels show OCR character-"
                          f"doubling — needs OCR-quality extraction")
        # (7)
        if oversize_cy / max(cy_present, 1) > 0.40:
            return True, (f"{sn}: {oversize_cy}/{cy_present} CY values are "
                          f"absurdly large (>1e9) — column-merge corruption")

    if sheets_with_data == 0:
        return True, "no sheet has a recognisable 'Particulars' column"

    return False, "extraction looks structurally intact"


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Post-mapper safety check: did the mapper produce a real report?
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Row layout in Financial Report (must mirror bs_pl_mapper.ROWS).
# Used for the BS imbalance check + YoY anomaly check below.
# Spare cells (F9 / F28 / F43 / F55) were removed — every value now
# lands in a named bucket. F19 (Lease Liab Current) is deprecated;
# lease values fold into F18 / F27 per Ind AS 116.
_NW_ROWS  = (5, 6, 7, 8)
_CL_ROWS  = (14, 15, 16, 17, 18, 19)
_NCL_ROWS = (24, 25, 26, 27)
_CA_ROWS  = (35, 36, 37, 38, 39, 40, 41, 42)
_NCA_ROWS = (48, 49, 50, 51, 52, 53, 54)

# Per-row YoY swing thresholds for anomaly detection. Each entry is
# (row_in_report_xlsx, label, max_swing_pct). A change beyond max_swing_pct
# from PY → CY raises a warning. Picked conservatively per line based on
# how volatile each metric realistically is for an Indian listed company:
#   - Share Capital almost never changes (rights issue is the rare cause)
#   - Revenue can swing 50% in a normal year (commodities, demerger)
#   - Net Profit can swing 200% (loss → profit recoveries, exceptional gains)
#   - Trade Payables / Receivables can swing 100% (working-capital cycle)
# Float entries are absolute fractions: 0.05 = 5%; 2.0 = 200%.
_YOY_THRESHOLDS = (
    (5,  "Share Capital",                       0.05),
    (8,  "Other Equity",                        0.50),
    (14, "Accounts Payable",                    1.00),
    (16, "Short term borrowings",               2.00),
    (24, "Long Term borrowings",                2.00),
    (37, "Inventory",                           1.00),
    (40, "Accounts Receivable",                 1.00),
    (48, "Fixed Assets",                        0.50),
    (51, "Capital Work-in-Progress",            5.00),  # CWIP commissioning
    (66, "Revenue",                             0.50),
    (69, "Employee benefits expense",           0.30),
    (71, "Depreciation",                        0.30),
    # Taxes moved from row 73 → 78 in the Schedule III re-layout
    # (rows 72/73 are now Other Expenses / Other Income, 74 is the
    # OE−OI formula, 76 Exceptional, 78 Taxes, 79 Net Profit).
    (78, "Taxes",                               2.00),
)
# Magnitudes below this (in Lakhs canonical) are ignored — a 500% swing
# on ₹2 lakh is noise, not an anomaly.
_YOY_NOISE_FLOOR = 50.0


def _sum_rows(ws, col: int, rows) -> float:
    total = 0.0
    for r in rows:
        v = ws.cell(row=r, column=col).value
        if isinstance(v, (int, float)):
            total += float(v)
    return total


def check_yoy_anomalies(report_xlsx_path: str) -> List[Tuple[str, str]]:
    """Check year-over-year swings on key lines and return a list of
    (severity, message) tuples for the WARNINGS sheet. Severity is one of:
        'info'    — within range, FYI only (not returned)
        'warn'    — exceeds threshold, reviewer should look
        'critical'— exceeds 5× threshold, almost certainly a bug

    For variant column pairs (CY,PY), e.g. Standalone CY at col 5 + PY at
    col 6, this compares per-row CY vs PY against `_YOY_THRESHOLDS`. Only
    fires when both values are above `_YOY_NOISE_FLOOR` to avoid noise on
    tiny line items.

    Catches things like:
      - Decimal-point bug (38,937 → 3,893.7 → 90% drop on Revenue)
      - Wrong-unit upload (Crore PDF mis-flagged as Lakh → 100× swing)
      - Mapper put a value in the wrong cell (massive swing on one line,
        opposite swing on another)
    """
    if openpyxl is None:
        return []
    path = Path(report_xlsx_path)
    if not path.exists():
        return []
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
    except Exception:
        return []
    if "Financial Report" not in wb.sheetnames:
        return []

    ws = wb["Financial Report"]

    # Find variant column pairs (CY, PY). Headers in row 2 contain the
    # variant + year label. Pair adjacent CY/PY columns of the same variant.
    headers: Dict[int, str] = {}
    for col in range(5, ws.max_column + 1):
        h = ws.cell(row=2, column=col).value
        if isinstance(h, str):
            headers[col] = h

    # A "pair" is (cy_col, py_col) where cy_col header contains "Current"
    # and the next column header contains "Prior" with the same variant.
    pairs: List[Tuple[int, int, str]] = []
    cols = sorted(headers.keys())
    i = 0
    while i < len(cols):
        c = cols[i]
        h = headers[c]
        if "current" in h.lower() and i + 1 < len(cols):
            next_c = cols[i + 1]
            next_h = headers[next_c]
            # Same variant prefix? Compare first word.
            v1 = h.split()[0] if h.split() else ""
            v2 = next_h.split()[0] if next_h.split() else ""
            if v1 == v2 and "prior" in next_h.lower():
                pairs.append((c, next_c, v1))
                i += 2
                continue
        i += 1

    findings: List[Tuple[str, str]] = []
    for cy_col, py_col, variant in pairs:
        for row, label, threshold in _YOY_THRESHOLDS:
            cy = ws.cell(row=row, column=cy_col).value
            py = ws.cell(row=row, column=py_col).value
            if not isinstance(cy, (int, float)) or not isinstance(py, (int, float)):
                continue
            cy_f = float(cy); py_f = float(py)
            if abs(cy_f) < _YOY_NOISE_FLOOR and abs(py_f) < _YOY_NOISE_FLOOR:
                continue
            if py_f == 0:
                # New line item this year; flag only if CY is large
                if abs(cy_f) > _YOY_NOISE_FLOOR * 10:
                    findings.append((
                        "warn",
                        f"YoY: {variant} {label} appeared CY={cy_f:.0f} "
                        f"with PY=0 (new line item or PY missed?)"
                    ))
                continue
            swing = (cy_f - py_f) / abs(py_f)
            if abs(swing) <= threshold:
                continue
            severity = "critical" if abs(swing) > threshold * 5 else "warn"
            sign = "+" if swing > 0 else ""
            findings.append((
                severity,
                f"YoY: {variant} {label} swung {sign}{swing*100:.0f}% "
                f"(CY={cy_f:.2f}, PY={py_f:.2f}, threshold ±{threshold*100:.0f}%)"
            ))
    return findings


def append_warnings_to_report(report_xlsx_path: str,
                                rows: List[Tuple[str, str, str]]) -> None:
    """Append rows to the WARNINGS sheet in an existing report xlsx.
    Each row is (variant, category, message). Creates the sheet if missing.
    Safe no-op when rows is empty."""
    if openpyxl is None or not rows:
        return
    path = Path(report_xlsx_path)
    if not path.exists():
        return
    try:
        wb = openpyxl.load_workbook(str(path))
    except Exception:
        return
    if "WARNINGS" not in wb.sheetnames:
        ws = wb.create_sheet("WARNINGS")
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 90
        ws["A1"] = "Variant"; ws["B1"] = "Category"; ws["C1"] = "Warning"
    else:
        ws = wb["WARNINGS"]
    next_row = ws.max_row + 1
    for variant, category, msg in rows:
        ws.cell(row=next_row, column=1, value=variant)
        ws.cell(row=next_row, column=2, value=category)
        ws.cell(row=next_row, column=3, value=msg)
        next_row += 1
    try:
        from extract_tables import _assert_not_protected
        _assert_not_protected(path)
    except ImportError:
        pass
    wb.save(str(path))


def is_report_broken(report_xlsx_path: str,
                      imbalance_pct_threshold: float = 0.01,
                      imbalance_abs_threshold: float = 50.0
                      ) -> Tuple[bool, str]:
    """Sanity-check the mapper's *_Report.xlsx output.

    Triggers (any one):
      1. No numeric data cells at all.
      2. Every variant column is all-zero.
      3. >85% of all data cells are exactly 0 (mostly-zero report).
      4. Any variant column has BS imbalance > max(1% of Total Assets,
         ₹50 lakh / 50 units in extract scale). Catches the case where
         the extractor missed a row, the mapper missed a row, or a
         fundamental section misclassification slipped through Claude's
         retry loop. Magnitudes <50 are ignored to avoid noise on
         genuinely tiny standalone PnL sheets.

    `imbalance_pct_threshold` and `imbalance_abs_threshold` are tunable so
    callers (or tests) can dial sensitivity without touching code.

    Returns (broken, reason).
    """
    if openpyxl is None:
        return False, "openpyxl unavailable"

    path = Path(report_xlsx_path)
    if not path.exists():
        return True, f"report xlsx missing: {path}"

    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
    except Exception as e:
        return True, f"failed to open report xlsx: {e}"

    if "Financial Report" not in wb.sheetnames:
        return True, "no 'Financial Report' sheet in mapper output"

    ws = wb["Financial Report"]

    # Existing zero-density checks (variant data starts at col E = 5)
    nonzero_cells = 0
    zero_cells = 0
    per_col_nonzero: Dict[int, int] = {}
    for col in range(5, ws.max_column + 1):
        per_col_nonzero[col] = 0
        for r in range(3, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if isinstance(v, (int, float)):
                if v == 0:
                    zero_cells += 1
                else:
                    nonzero_cells += 1
                    per_col_nonzero[col] += 1

    total = nonzero_cells + zero_cells
    if total == 0:
        return True, "report has no numeric data cells at all"

    if nonzero_cells == 0:
        return True, f"report is 100% zeros across {total} cells"

    empty_cols = [col for col, n in per_col_nonzero.items() if n == 0]
    if empty_cols and len(empty_cols) == len(per_col_nonzero):
        return True, "every variant column is all-zero"

    if zero_cells / total > 0.85:
        return True, (f"{zero_cells}/{total} data cells are 0 (>85%) — "
                       f"mapper produced a near-empty report")

    # NEW: BS imbalance check per variant column. Catches Finolex-style
    # failures where the extract dropped rows and Claude couldn't recover
    # — Total L+E reads ₹5,027 cr but Total Assets reads ₹2,655 cr, a
    # 47% gap. The mapper's own balance check flagged it but the warning
    # never blocked the pipeline. Now it triggers the Claude PDF fallback.
    for col in range(5, ws.max_column + 1):
        if per_col_nonzero.get(col, 0) == 0:
            continue
        nw  = _sum_rows(ws, col, _NW_ROWS)
        cl  = _sum_rows(ws, col, _CL_ROWS)
        ncl = _sum_rows(ws, col, _NCL_ROWS)
        ca  = _sum_rows(ws, col, _CA_ROWS)
        nca = _sum_rows(ws, col, _NCA_ROWS)
        assets = ca + nca
        liab_eq = nw + cl + ncl
        if assets == 0 and liab_eq == 0:
            continue
        diff = round(assets - liab_eq, 2)
        if abs(diff) <= imbalance_abs_threshold:
            continue
        scale = max(abs(assets), abs(liab_eq))
        pct = abs(diff) / scale if scale else 1.0
        if pct >= imbalance_pct_threshold:
            header = ws.cell(row=2, column=col).value or f"col{col}"
            header_clean = str(header).replace("\n", " / ")
            return True, (f"BS imbalance in '{header_clean}': "
                           f"Assets={assets:.2f}, Liab+Equity={liab_eq:.2f}, "
                           f"gap={diff:+.2f} ({pct*100:.1f}% of scale). "
                           f"Likely a missed row in extraction or "
                           f"misclassified section.")

    return False, "report looks populated"


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Claude PDF fallback
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

DEFAULT_MODEL = "claude-sonnet-4-6"

# Statement IDs Claude will return in the `sheets[].kind` enum. Mirrors the
# keys of `STMT_NAMES` in extract_tables.py.
_STMT_ENUM = list(STMT_NAMES.keys())

# Tool schema: Claude returns a list of sheets, each a list of rows. Rows
# preserve the source structure: particulars, optional note, CY value,
# PY value. Numeric values are floats (negatives for parentheses). Empty
# cells become null.
_TOOL = {
    "name": "submit_extracted_statements",
    "description": (
        "Submit the structured BS / P&L tables extracted from the PDF. "
        "Return one entry in `sheets` per distinct financial statement "
        "(consolidated BS, standalone BS, consolidated P&L, standalone P&L). "
        "Keep the source order of rows. Use null for empty cells, not 0. "
        "Negative numbers (figures in parentheses) must be returned as "
        "negative floats, not strings."
    ),
    "input_schema": {
        "type": "object",
        "properties": {
            "company": {
                "type": "string",
                "description": "Company name as printed on the statement."
            },
            "currency_unit": {
                "type": "string",
                "enum": ["lakh", "crore", "million", "billion",
                          "thousand", "rupee"],
                "description": ("Canonical unit token for the values in "
                                "this PDF. Read the header — '(Rs in Crore)' "
                                "→ 'crore'; '(INR Lakhs)' → 'lakh'; '(USD "
                                "Million)' → 'million'. Default to 'lakh' if "
                                "no unit is declared on Indian Schedule III "
                                "statements.")
            },
            "sheets": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "kind": {
                            "type": "string",
                            "enum": _STMT_ENUM,
                            "description": ("Statement type. Use the "
                                            "consolidated_* / standalone_* "
                                            "variants when explicitly "
                                            "labelled; otherwise generic_*.")
                        },
                        "current_year_label": {
                            "type": "string",
                            "description": ("Header text for the current-year "
                                            "column, e.g. 'March 31, 2023'.")
                        },
                        "prior_year_label": {
                            "type": "string",
                            "description": "Header text for the prior-year column."
                        },
                        "rows": {
                            "type": "array",
                            "items": {
                                "type": "object",
                                "properties": {
                                    "particulars": {"type": "string"},
                                    "note": {"type": ["string", "null"]},
                                    "current_year": {"type": ["number", "null"]},
                                    "prior_year": {"type": ["number", "null"]},
                                },
                                "required": ["particulars", "note",
                                             "current_year", "prior_year"],
                            },
                        },
                    },
                    "required": ["kind", "current_year_label",
                                 "prior_year_label", "rows"],
                },
            },
        },
        "required": ["company", "currency_unit", "sheets"],
    },
}

_SYSTEM = (
    "You are a careful financial-statement parser. The user uploads an "
    "Indian Ind-AS / Old-GAAP balance sheet and statement of profit & loss "
    "(possibly both consolidated and standalone variants in the same PDF). "
    "Your job is to extract every line item with its current-year and "
    "prior-year value into structured rows.\n\n"
    "RULES:\n"
    "1. ALWAYS call the submit_extracted_statements tool. Never reply in "
    "free text.\n"
    "2. One entry in `sheets` per distinct statement (so a PDF with both a "
    "consolidated BS and a standalone BS gives 2 BS entries).\n"
    "3. Preserve source row order. Include section headers (\"ASSETS\", "
    "\"Non-current assets\") as rows with null values — they anchor the "
    "downstream mapper.\n"
    "4. Numbers in parentheses are negative — return -123.45, not "
    "\"(123.45)\".\n"
    "5. Empty cells must be null, not 0 and not \"-\".\n"
    "6. Drop page-design watermark text and tab/sidebar tokens "
    "(\"Statutory Reports\", \"Creating Value\", page numbers, etc.) — "
    "they are NOT data.\n"
    "7. Fix obvious OCR character-doubling (e.g. \"IInnccoommee\" -> "
    "\"Income\"). Output the cleaned label.\n"
    "8. Do not invent rows. If a label has no values on either side, still "
    "emit it as a header row (both years null) so downstream knows the "
    "section structure."
)


def _model_for(cfg: Dict[str, Any], override: Optional[str]) -> str:
    if override:
        return override
    m = (cfg.get("claude_model") or "").strip()
    return m or DEFAULT_MODEL


def _api_key_for(cfg: Dict[str, Any], override: Optional[str]) -> str:
    if override:
        return override.strip()
    return ((cfg.get("anthropic_api_key") or "").strip()
            or os.environ.get("ANTHROPIC_API_KEY", "").strip())


def _load_cfg() -> Dict[str, Any]:
    try:
        from config import load_config
        return load_config()
    except Exception:
        return {}


def _call_claude(pdf_bytes: bytes, model: str, api_key: str) -> Optional[Dict[str, Any]]:
    try:
        from anthropic import Anthropic
    except ImportError:
        print("    [claude_pdf_extractor] anthropic SDK not installed. "
              "Run: pip install anthropic")
        return None

    client = Anthropic(api_key=api_key)
    pdf_b64 = base64.standard_b64encode(pdf_bytes).decode("ascii")

    try:
        resp = client.messages.create(
            model=model,
            max_tokens=16000,
            system=_SYSTEM,
            tools=[_TOOL],
            tool_choice={"type": "tool", "name": _TOOL["name"]},
            messages=[{
                "role": "user",
                "content": [
                    {
                        "type": "document",
                        "source": {
                            "type": "base64",
                            "media_type": "application/pdf",
                            "data": pdf_b64,
                        },
                        # Cache control: subsequent retries / multi-PDF
                        # batches reuse the parsed PDF on Anthropic's side.
                        "cache_control": {"type": "ephemeral"},
                    },
                    {
                        "type": "text",
                        "text": ("Extract every BS / P&L statement in this "
                                  "PDF using the submit_extracted_statements "
                                  "tool."),
                    },
                ],
            }],
        )
    except Exception as e:
        print(f"    [claude_pdf_extractor] API call failed: {type(e).__name__}: {e}")
        return None

    # Extract tool_use block
    for block in resp.content:
        if getattr(block, "type", None) == "tool_use" \
                and getattr(block, "name", None) == _TOOL["name"]:
            return block.input  # type: ignore[attr-defined]
    print("    [claude_pdf_extractor] model did not return a tool_use block "
          "(stop_reason=" + str(getattr(resp, "stop_reason", "?")) + ")")
    return None


def _payload_to_rows(sheet_payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Convert Claude's row dicts into the shape `write_sheet` expects:
    {particulars, note, cy, py, cy_raw, py_raw}.
    Empty values become '' (not None) to match parse_value() conventions."""
    out = []
    for r in sheet_payload.get("rows", []):
        part = (r.get("particulars") or "").strip()
        note = (r.get("note") or "").strip() if r.get("note") is not None else ""
        cy = r.get("current_year")
        py = r.get("prior_year")
        # write_sheet treats numeric vs string differently — preserve None as ''.
        cy_val = cy if isinstance(cy, (int, float)) else ''
        py_val = py if isinstance(py, (int, float)) else ''
        out.append({
            "particulars": part,
            "note": note,
            "cy": cy_val,
            "py": py_val,
            "cy_raw": "" if cy is None else str(cy),
            "py_raw": "" if py is None else str(py),
        })
    return out


def extract_with_claude(pdf_path: str, output_dir: str,
                         api_key: Optional[str] = None,
                         model: Optional[str] = None) -> Optional[str]:
    """Extract statements via Claude PDF input and write a *_extracted.xlsx
    in `output_dir`. Returns the file path on success, None on failure.

    On failure (no key, SDK missing, API error, malformed response) returns
    None so the caller can keep the rule-based extracted xlsx and continue."""
    if openpyxl is None:
        print("    [claude_pdf_extractor] openpyxl not available")
        return None

    pdf_path_p = Path(pdf_path)
    if not pdf_path_p.exists():
        print(f"    [claude_pdf_extractor] PDF not found: {pdf_path}")
        return None

    cfg = _load_cfg()
    key = _api_key_for(cfg, api_key)
    if not key:
        print("    [claude_pdf_extractor] no anthropic_api_key configured "
              "— skipping Claude fallback")
        return None
    used_model = _model_for(cfg, model)

    # Anthropic accepts PDFs up to 32MB / 100 pages on Sonnet 4.6. Bigger
    # documents are rare for BS/PL extracts (the page_detector trims to
    # ~10 pages) so we just pass the bytes through.
    pdf_bytes = pdf_path_p.read_bytes()
    size_mb = len(pdf_bytes) / (1024 * 1024)
    print(f"    [claude_pdf_extractor] sending {pdf_path_p.name} "
          f"({size_mb:.2f} MB) to {used_model}")

    payload = _call_claude(pdf_bytes, used_model, key)
    if not payload or not isinstance(payload, dict):
        return None

    sheets = payload.get("sheets") or []
    if not sheets:
        print("    [claude_pdf_extractor] model returned 0 sheets")
        return None

    company = (payload.get("company") or "").strip()
    source_unit = (payload.get("currency_unit") or "lakh").strip().lower()
    if source_unit not in ("lakh", "crore", "million",
                            "billion", "thousand", "rupee"):
        source_unit = "lakh"
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{pdf_path_p.stem}_extracted.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    used_names: set[str] = set()
    written = 0
    for sp in sheets:
        kind = sp.get("kind") or "generic_bs"
        rows = _payload_to_rows(sp)
        if not rows:
            continue
        # Apply unit normalization: Claude returns values as published in
        # the PDF; the downstream mapper expects Lakhs.
        rows, factor = normalize_rows_to_lakhs(rows, source_unit)
        if factor != 1.0 and written == 0:
            print(f"    [claude_pdf_extractor] unit='{source_unit}' "
                   f"→ scaling values by {factor:g} to Lakhs")
        sheet_label = STMT_NAMES.get(kind, kind)
        sn = sheet_label[:31]
        ctr = 2
        while sn in used_names:
            sn = f"{sheet_label[:28]}_{ctr}"
            ctr += 1
        used_names.add(sn)
        ws = wb.create_sheet(title=sn)
        write_sheet(
            ws, rows,
            sp.get("current_year_label") or "Current Year",
            sp.get("prior_year_label") or "Prior Year",
            company, sheet_label,
            source_unit=source_unit,
        )
        print(f"    [claude_pdf_extractor]   '{sn}': {len(rows)} rows")
        written += 1

    if written == 0:
        return None

    # Defensive guard: never overwrite hand-curated templates.
    try:
        from extract_tables import _assert_not_protected
        _assert_not_protected(out_path)
    except ImportError:
        pass

    wb.save(str(out_path))
    print(f"    [claude_pdf_extractor] wrote {out_path}")
    return str(out_path)
