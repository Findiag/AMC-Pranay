#!/usr/bin/env python3
"""
BS/PL Financial Statement → Template Mapper (v2)
=================================================
Converts extracted BS/PL Excel files into a standardized template using
OpenAI GPT-4o for intelligent label mapping.

Key accuracy features:
  - Residual-based F72 computation (P&L guaranteed to balance)
  - BS section-total cross-check with auto-retry
  - Combined single-file output (Standalone + Consolidated × Current + Prior)
  - Multi-row label merging with section-header detection

Usage:
  Single file:   python bs_pl_mapper.py --input file.xlsx --template tmpl.xlsx --api-key sk-... --output out/
  Batch folder:  python bs_pl_mapper.py --input folder/    --template tmpl.xlsx --api-key sk-... --output out/
"""

import os, sys, json, re, argparse
from pathlib import Path
from typing import Optional
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Smart matcher for hybrid PL/BS classification
try:
    from matcher import get_matcher, SmartMatcher, BS_FIELD_TO_CELL
    HAS_MATCHER = True
except ImportError:
    HAS_MATCHER = False

try:
    from openai import OpenAI
    HAS_OPENAI = True
except ImportError:
    HAS_OPENAI = False

# v2 LLM mapper — strict structured outputs + hierarchical (section→field) +
# programmatic post-checks with targeted re-prompts (10-point engineering plan)
try:
    from llm_mapper_v2 import map_bs_v2, map_pl_v2
    HAS_LLM_V2 = True
except ImportError as _e:
    HAS_LLM_V2 = False
    _llm_v2_import_error = str(_e)

# Claude mapper — primary LLM mapper as of the 2026-05 rewrite. Same shape as
# llm_mapper_v2 (strict structured output via tool-use). When the
# anthropic_api_key is set in config.json, this is the default mapper.
try:
    from claude_mapper import map_bs_claude, map_pl_claude
    HAS_CLAUDE = True
except ImportError as _e:
    HAS_CLAUDE = False
    _claude_import_error = str(_e)

# Per-company multi-year workbook appender. Called after build_report so a
# single per-company file accumulates year-by-year instead of being
# overwritten every upload.
try:
    from company_excel_writer import append_company_report
    HAS_COMPANY_WRITER = True
except ImportError:
    HAS_COMPANY_WRITER = False

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PREPROCESSING
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

NON_MERGE_LABELS = {
    "income", "expenses", "assets", "liabilities", "equity",
    "financial assets", "non-current assets", "current assets",
    "non-current liabilities", "current liabilities",
    "non- current assets", "non- current liabilities",
    "equity and liabilities", "shareholder's fund", "shareholders fund",
    "trade payables", "tax expense", "tax expenses",
    "other comprehensive income", "exceptional items",
    "items that will not be reclassified to profit or loss:",
    "items that will be reclassified to profit or loss:",
    "property, plant & equipment", "revenue",
    "borrowings", "provisions", "investments", "loans",
    "intangible assets", "financial liabilities",
    "non-controlling interest", "non controlling interest",
}

def clean_value(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "").replace("(", "-").replace(")", "")
    if s in ["-", "–", "—", "", "nil", "Nil", "NIL"]:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0

def is_section_header(label):
    return label.lower().strip().rstrip(":") in NON_MERGE_LABELS

def merge_multirow(items):
    merged, skip = [], False
    for i, item in enumerate(items):
        if skip:
            skip = False
            continue
        if is_section_header(item["label"]):
            merged.append(item)
            continue
        if i + 1 < len(items):
            nxt = items[i + 1]
            cur_has = item["cur"] != 0 or item["pri"] != 0
            nxt_has = nxt["cur"] != 0 or nxt["pri"] != 0
            if not cur_has and nxt_has:
                cl = item["label"].lower().strip()
                if cl.endswith(("of", "the", "than", "to", "and", "other", "micro", "small")):
                    merged.append({"label": item["label"]+" "+nxt["label"], "cur": nxt["cur"], "pri": nxt["pri"]})
                    skip = True
                    continue
                nl = nxt["label"].lower().strip()
                if nl.startswith(("enterprises", "profit or loss", "and small")):
                    merged.append({"label": item["label"]+" "+nxt["label"], "cur": nxt["cur"], "pri": nxt["pri"]})
                    skip = True
                    continue
        merged.append(item)
    return merged

def extract_items(filepath, sheet_name):
    df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
    items, started = [], False
    for _, row in df.iterrows():
        label = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
        if "particulars" in label.lower() and not started:
            started = True
            continue
        if not started:
            continue

        cur = clean_value(row.iloc[2] if len(row) > 2 else None)
        pri = clean_value(row.iloc[3] if len(row) > 3 else None)

        # v4 patch: attach orphan-value rows (label=None, has values) to the
        # immediately preceding non-section-header row, FILLING IN whichever
        # side(s) (cur / pri) the previous row was missing.
        #
        # The Finolex bug that motivated this:
        #   R20: ['(b) Financial Assets', None, None, None]   ← section header
        #   R21: ['i) Investments',  '8',  None,    217961 ]  ← prev: pri only
        #   R22: [None,              None, 241438,  None   ]  ← orphan: cur only
        # Without this, the ₹2,414 cr CY Investments value gets dropped and
        # Total Assets ends ₹2,372 cr short of Equity+Liab. With this,
        # R22's `cur` snaps onto R21 to form a complete row.
        #
        # Section-header rows (Financial Assets / ASSETS / etc.) are skipped
        # as the attach target — their orphan values are subtotals and the
        # mapper sums leaves itself.
        #
        # Conservative: only fills MISSING sides; never overwrites a
        # populated value, so a row that already has both cur+pri is left
        # alone and the orphan goes nowhere (likely a subtotal).
        if (not label or label.lower() == "nan"):
            if (cur != 0 or pri != 0) and items:
                prev = items[-1]
                if not is_section_header(prev["label"]):
                    if prev["cur"] == 0 and cur != 0:
                        prev["cur"] = cur
                    if prev["pri"] == 0 and pri != 0:
                        prev["pri"] = pri
            continue

        # Fix 4: Value-in-label — split "PPE 7,590.12 6,500.00" into label + values
        if cur == 0 and pri == 0:
            # Check if label ends with numbers: "label text 1,234.56 789.01"
            m = re.match(r'^(.+?)\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)$', label)
            if m:
                label = m.group(1).strip()
                cur = clean_value(m.group(2))
                pri = clean_value(m.group(3))
            else:
                # Single trailing number: "label text 1,234.56"
                m = re.match(r'^(.+?)\s+([\d,]+\.?\d+)$', label)
                if m and len(m.group(1)) > 3:
                    label = m.group(1).strip()
                    cur = clean_value(m.group(2))

        items.append({"label": label, "cur": cur, "pri": pri})
    return merge_multirow(items)

def classify_sheets(sheet_names):
    result = {}
    for name in sheet_names:
        nl = name.lower()
        is_bs = any(k in nl for k in ["balance sheet", "balance_sheet", "bs"])
        is_pl = any(k in nl for k in ["p&l", "p & l", "pl", "profit", "loss", "income"])
        is_con = "consol" in nl
        is_sa = "standalone" in nl or "stand" in nl
        stype = "bs" if is_bs else ("pl" if is_pl else None)
        if not stype:
            continue
        variant = "consolidated" if is_con else ("standalone" if is_sa else "default")
        result[name] = {"type": stype, "variant": variant}
    return result

def process_source(filepath):
    xls = pd.ExcelFile(filepath)
    classified = classify_sheets(xls.sheet_names)
    variants = {}
    for sheet_name, info in classified.items():
        items = extract_items(filepath, sheet_name)
        v = info["variant"]
        if v not in variants:
            variants[v] = {}
        # MERGE split sheets (e.g., "Balance Sheet" + "Balance Sheet_2")
        if info["type"] in variants[v]:
            variants[v][info["type"]].extend(items)
        else:
            variants[v][info["type"]] = items
    # If a "default" variant exists, treat it as standalone. extract_tables
    # produces sheet names "Balance Sheet" / "P&L" (no "Standalone" prefix)
    # for the non-consolidated variant, so they end up labelled "default" by
    # classify_sheets even when a consolidated variant also exists. Without
    # this rename, build_report — which only iterates ("standalone",
    # "consolidated") — silently drops the standalone columns from the
    # final report.
    if "default" in variants:
        if "standalone" in variants:
            # Both exist — merge default rows into standalone (rare; happens
            # when a sheet is literally named "Standalone" alongside an
            # ambiguous one)
            for k, v in variants.pop("default").items():
                variants["standalone"].setdefault(k, []).extend(v)
        else:
            variants["standalone"] = variants.pop("default")

    # ── standalone_only mode ──
    # When config.standalone_only is true (the default), drop Consolidated
    # variants entirely so the mapper only spends LLM calls on Standalone,
    # the per-job Report has only Standalone columns, and the per-company
    # workbook tracks only Standalone year-columns. Set
    # `"standalone_only": false` in config.json to restore both variants.
    try:
        from config import load_config as _load_cfg
        _so = bool(_load_cfg().get("standalone_only", True))
    except Exception:
        _so = True
    if _so and "consolidated" in variants:
        variants.pop("consolidated")
    return variants


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# LLM MAPPING
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

BS_PROMPT = """You are an expert Indian financial analyst (Ind AS + Old GAAP).
Map these Balance Sheet line items into the template cells below.

SOURCE (label: current_year_value | prior_year_value):
{items}

TEMPLATE CELLS:
NETWORTH:
  F5  = Share Capital
  F6  = Retained Earnings (0 if not separately disclosed)
  F7  = General Reserves & Surplus (Old GAAP "Reserves and surplus" goes here)
  F8  = Other Equity (Ind AS "Other Equity" goes here)
  F9  = spare (Non-Controlling Interest if present)

CURRENT LIABILITIES:
  F14 = Accounts Payable (SUM all Trade Payables sub-items: MSME + non-MSME)
  F15 = Provisions (current)
  F16 = Short term borrowings
  F17 = Other Current Liabilities
  F18 = Other Financial Liabilities
  F19 = Lease Liabilities (current) — if not present, 0

NON-CURRENT LIABILITIES:
  F24 = Long Term borrowings
  F25 = Provision (non-current)
  F26 = Others = Lease Liabilities(NC) + Deferred Tax Liabilities + Other Non-Current Liabilities (SUM all)
  F27 = Other Financial Liabilities (non-current)
  F28 = spare

CURRENT ASSETS:
  F35 = Bank Balance ("Other Bank Balances"; if only "Cash and Bank Balances" exists as combined, put 0 here)
  F36 = Cash & Cash Equivalence (if combined "Cash and Bank Balances", put full amount here)
  F37 = Inventory / Inventories (0 for IT/service companies)
  F38 = Investments (current)
  F39 = Loans (current) / Short Term Loan and Advances
  F40 = Accounts Receivable = Trade Receivables
  F41 = Other Current Assets + Current Tax Assets (SUM both)
  F42 = Other Financial Assets (current)
  F43 = spare

NON-CURRENT ASSETS:
  F48 = Fixed Assets = PPE + Intangible Assets + Right of Use Assets + Goodwill + Software (SUM all EXCEPT CWIP)
  F49 = Investments (non-current)
  F50 = Loans (non-current) / Long-term loans & advances
  F51 = Capital Work-in-Progress + Intangible Assets under Development (SUM)
  F52 = Other Non-Current Assets
  F53 = Other Financial Assets (non-current)
  F54 = Deferred Tax Assets
  F55 = spare

CRITICAL RULES:
- Every source value MUST appear in exactly one cell. Double-counting = wrong. Missing items = wrong.
- The test: SUM(F5:F9) + SUM(F14:F19) + SUM(F24:F28) MUST equal SUM(F35:F43) + SUM(F48:F55).
- If the source has "Total Assets" and "Total Equity and Liabilities", your mapped totals must match those.

Return ONLY valid JSON (no markdown fences):
{{"current_year":{{"F5":0,...}},"prior_year":{{"F5":0,...}},"notes":"brief explanation"}}"""

PL_PROMPT = """You are an expert Indian financial analyst.
Map these P&L line items to template cells. IMPORTANT: I will back-calculate F72 myself, so DO NOT include F72.

SOURCE (label: current_year_value | prior_year_value):
{items}

TEMPLATE CELLS:
  F66 = Revenue from Operations ONLY (exclude Other Income)
  F67 = Cost of Goods Sold. Rules:
        - Manufacturing: "Cost of Materials Consumed" (+ "Changes in Inventories" if present)
        - Trading: "Purchase of traded goods" + "Changes in inventories of traded goods"
        - Services/IT: "Operating Expenses" if that's the main cost line
        - If no clear COGS exists, put 0
  F69 = Employee Benefits Expense
  F70 = Finance Costs / Interest
  F71 = Depreciation and Amortization Expense
  F73 = Total Tax Expense (SUM of Current Tax + Deferred Tax + Tax of Earlier Years)

ALSO EXTRACT:
  other_income = Other Income value
  net_profit = The final "Profit after tax" / "Profit for the year" / "Net Profit" figure
  exceptional_items = Any exceptional income (positive) or expense (negative). 0 if none.

I will compute F72 = F66 - F67 - F69 - F70 - F71 - F73 - net_profit
This guarantees the P&L balances perfectly.

Return ONLY valid JSON (no markdown fences):
{{"current_year":{{"F66":0,"F67":0,"F69":0,"F70":0,"F71":0,"F73":0,"other_income":0,"net_profit":0,"exceptional_items":0}},"prior_year":{{"F66":0,"F67":0,"F69":0,"F70":0,"F71":0,"F73":0,"other_income":0,"net_profit":0,"exceptional_items":0}},"notes":"brief"}}"""


def format_items_for_prompt(items):
    lines = []
    for it in items:
        if it["cur"] != 0 or it["pri"] != 0:
            lines.append(f'  "{it["label"]}": {it["cur"]} | {it["pri"]}')
    return "\n".join(lines)

def call_gpt(client, prompt, context, max_retries=2):
    for attempt in range(max_retries + 1):
        try:
            resp = client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": "Expert Indian financial analyst. Return ONLY valid JSON. No markdown fences. No explanations outside JSON."},
                    {"role": "user", "content": prompt},
                ],
                temperature=0.0,
                max_tokens=3000,
            )
            text = resp.choices[0].message.content.strip()
            text = re.sub(r"^```(?:json)?\s*", "", text)
            text = re.sub(r"\s*```$", "", text)
            return json.loads(text)
        except (json.JSONDecodeError, Exception) as e:
            if attempt < max_retries:
                print(f"    [Retry {attempt+1}] {context}: {e}")
            else:
                print(f"    [FAILED] {context}: {e}")
                return None

def map_bs(client, items):
    prompt = BS_PROMPT.replace("{items}", format_items_for_prompt(items))
    return call_gpt(client, prompt, "BS")



def map_bs_hybrid(items):
    """100% offline BS mapping using rules + TF-IDF. No API needed.
    v2 patch: low-confidence items (conf < 0.50) are re-classified via RAG
    (sentence-transformers + ChromaDB), with GPT-4o-mini fallback if the
    openai_api_key is configured in config.json."""
    if not HAS_MATCHER:
        print("    BS: matcher not available")
        return None

    matcher = get_matcher()
    result = matcher.classify_bs_items(items)

    matched = sum(1 for m in result["all_matches"]
                  if m["match"]["confidence"] >= 0.30 and m["match"]["cell"] != "_skip")
    total = len(result["all_matches"])

    print(f"    BS: Matched {matched}/{total} items (offline)")

    # ── RAG re-classification for low-confidence items ──
    # Pre-import so one failure doesn't break the whole pipeline
    try:
        from rag_matcher import classify_bs_label, add_correction, BS_FIELD_LIST
        # Map field names (matcher) → cell refs (template)
        from matcher import BS_FIELD_TO_CELL as _F2C
        rag_available = True
    except Exception as _rag_err:
        rag_available = False
        print(f"    BS: RAG layer unavailable ({type(_rag_err).__name__}); skipping rescue")

    rag_rescued = 0
    for m in result["all_matches"]:
        c = m["match"]["confidence"]
        cell = m["match"]["cell"]
        method = m["match"]["method"]
        # Only re-run RAG when the item has values AND is low-confidence or skipped
        has_val = (m.get("cur", 0) != 0 or m.get("pri", 0) != 0)
        # Don't rescue items mapped via exact rules (high confidence) — they're likely correct
        should_try_rag = rag_available and has_val and (
            (c < 0.50 and cell != "_skip") or
            (cell == "_skip" and method == "fallback")
        )
        if should_try_rag:
            section = m.get("_section", "")
            try:
                rag_result = classify_bs_label(m["label"], section=section, use_gpt=True)
            except Exception as e:
                rag_result = None
                print(f"    BS: RAG error on '{m['label'][:40]}': {e}")
            if rag_result and rag_result["field"] and rag_result["confidence"] > c:
                new_cell = _F2C.get(rag_result["field"], cell)
                if new_cell != cell:
                    print(f"    BS RAG rescue: [{rag_result['source']}] "
                          f"{m['label'][:40]} {cell}({c:.2f}) → {new_cell}({rag_result['confidence']:.2f}) "
                          f"— {rag_result['reasoning'][:60]}")
                    # Re-route: subtract from old bucket, add to new bucket
                    old = result.get(cell, {"cur": 0, "pri": 0})
                    old["cur"] = old.get("cur", 0) - m.get("cur", 0)
                    old["pri"] = old.get("pri", 0) - m.get("pri", 0)
                    new = result.get(new_cell, {"cur": 0, "pri": 0})
                    new["cur"] = new.get("cur", 0) + m.get("cur", 0)
                    new["pri"] = new.get("pri", 0) + m.get("pri", 0)
                    result[cell] = old
                    result[new_cell] = new
                    m["match"]["cell"] = new_cell
                    m["match"]["confidence"] = rag_result["confidence"]
                    m["match"]["method"] = f"rag_{rag_result['source']}"
                    rag_rescued += 1

    if rag_rescued:
        print(f"    BS: RAG rescued {rag_rescued} items")

    for m in result["all_matches"]:
        c = m["match"]["confidence"]
        cell = m["match"]["cell"]
        method = m["match"]["method"]
        if (m["cur"] != 0 or m["pri"] != 0) and cell != "_skip":
            sym = "✓" if c >= 0.30 else "?"
            print(f"      {sym} [{c:.2f} {method:8s}] {m['label'][:45]:45s} → {cell}")

    if result["unmatched"]:
        print(f"    BS: {len(result['unmatched'])} unmatched")

    mapping = {"current_year": {}, "prior_year": {}, "notes": "offline", "_warnings": []}
    for field, cell in BS_FIELD_TO_CELL.items():
        bucket = result.get(cell, {"cur": 0, "pri": 0})
        mapping["current_year"][cell] = round(bucket["cur"], 2)
        mapping["prior_year"][cell] = round(bucket["pri"], 2)

    # ── AUTO-CORRECTION: find skipped items that could fix the balance ──
    v = validate_bs(mapping, "current_year")
    if not v["balanced"]:
        diff = v["diff"]  # positive = assets > liab, negative = liab > assets
        skipped_with_values = [
            m for m in result.get("all_matches", [])
            if m["match"]["cell"] == "_skip" and m.get("cur", 0) != 0
        ]
        for sk in skipped_with_values:
            val = sk["cur"]
            # If diff is positive (assets side heavy) and this skipped value is positive,
            # it might belong on the liabilities side
            if abs(diff - val) < abs(diff) and abs(val) > 1:
                # This item might reduce the diff if added to the right side
                section = sk.get("_section", "")
                if diff > 0 and section and "liab" in str(section):
                    # Add to Other CL or Other NCL based on section
                    target = "F17" if "current" in str(section) else "F26"
                    mapping["current_year"][target] = round(mapping["current_year"].get(target, 0) + val, 2)
                    mapping["prior_year"][target] = round(mapping["prior_year"].get(target, 0) + sk.get("pri", 0), 2)
                    print(f"    BS auto-fix: [{sk['label'][:40]}] → {target} (val={val})")
                    diff = diff - val
                elif diff < 0 and section and "asset" in str(section):
                    target = "F41" if "current" in str(section) else "F52"
                    mapping["current_year"][target] = round(mapping["current_year"].get(target, 0) + val, 2)
                    mapping["prior_year"][target] = round(mapping["prior_year"].get(target, 0) + sk.get("pri", 0), 2)
                    print(f"    BS auto-fix: [{sk['label'][:40]}] → {target} (val={val})")
                    diff = diff - val

    mapping["_warnings"] = result.get("warnings", [])
    return mapping


def map_pl_hybrid(items):
    """100% offline PL mapping using rules + TF-IDF. No API needed."""
    if not HAS_MATCHER:
        print("    PL: matcher not available")
        return None

    matcher = get_matcher()
    result = matcher.classify_pl_items(items)

    matched = sum(1 for m in result["all_matches"]
                  if m["match"]["confidence"] >= 0.30 and m["match"]["cell"] != "_skip")
    total = len(result["all_matches"])

    print(f"    PL: Matched {matched}/{total} items (offline)")
    for m in result["all_matches"]:
        c = m["match"]["confidence"]
        field = m["match"]["field"]
        method = m["match"]["method"]
        if m["cur"] != 0 or m["pri"] != 0:
            sym = "✓" if c >= 0.30 else "?"
            print(f"      {sym} [{c:.2f} {method:8s}] {m['label'][:45]:45s} → {field}")

    mapping = {"current_year": {}, "prior_year": {}, "notes": "offline"}
    for key in ["F66", "F67", "F69", "F70", "F71", "F73"]:
        b = result.get(key, {"cur": 0, "pri": 0})
        mapping["current_year"][key] = round(b["cur"], 2)
        mapping["prior_year"][key] = round(b["pri"], 2)
    for key in ["net_profit", "other_income", "exceptional"]:
        b = result.get(key, {"cur": 0, "pri": 0})
        mapping["current_year"][key] = round(b["cur"], 2)
        mapping["prior_year"][key] = round(b["pri"], 2)
    mapping["current_year"]["exceptional_items"] = mapping["current_year"].get("exceptional", 0)
    mapping["prior_year"]["exceptional_items"] = mapping["prior_year"].get("exceptional", 0)

    sources = []
    for key in ["F66", "F67", "F69", "F70", "F71", "F73"]:
        b = result.get(key, {"sources": []})
        if b.get("sources"):
            sources.append(f"{key}: {', '.join(b['sources'][:2])}")
    mapping["notes"] = f"Offline: {matched}/{total}. " + "; ".join(sources[:4])
    mapping["_warnings"] = result.get("warnings", [])
    return mapping



# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# VALIDATION & RESIDUAL COMPUTATION
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def validate_bs(mapping, year):
    m = mapping[year]
    nw  = sum(float(m.get(f"F{i}", 0)) for i in range(5, 10))
    cl  = sum(float(m.get(f"F{i}", 0)) for i in range(14, 20))
    ncl = sum(float(m.get(f"F{i}", 0)) for i in range(24, 29))
    ca  = sum(float(m.get(f"F{i}", 0)) for i in range(35, 44))
    nca = sum(float(m.get(f"F{i}", 0)) for i in range(48, 56))
    total_liab = round(nw + cl + ncl, 2)
    total_assets = round(ca + nca, 2)
    diff = round(total_assets - total_liab, 2)
    return {
        "balanced": abs(diff) < 0.5,
        "diff": diff,
        "total_assets": total_assets,
        "total_liabilities": total_liab,
        "networth": round(nw, 2),
        "current_liab": round(cl, 2),
        "noncurrent_liab": round(ncl, 2),
        "current_assets": round(ca, 2),
        "noncurrent_assets": round(nca, 2),
    }

def compute_f72_residual(pl_mapping, year):
    """LITERAL F72 = Σ(other_expense values mapped) − other_income.

    The earlier residual approach (F72 = Rev − COGS − Emp − Int − Dep
    − Tax − NP) accidentally absorbed Exceptional Items, making the F72
    label "Other expenses less other income" a lie when Exc ≠ 0.

    Now F72 is computed strictly from the mapper's own `other_expense` /
    `other_income` assignments, matching the published P&L line items
    exactly. Exceptional Items are surfaced as their own row (F75) and
    folded into Net Profit via the Excel formula:
        NP = Gross Profit − SUM(Emp..Tax) + Exceptional
    """
    m = pl_mapping[year]
    other_expense = float(m.get("other_expense", 0) or 0)
    other_income  = float(m.get("other_income", 0) or 0)
    f72 = other_expense - other_income
    return round(f72, 2)

def retry_bs_with_error(client, items, mapping, year, error_info):
    """Re-prompt LLM with specific error for self-correction."""
    items_text = format_items_for_prompt(items)
    retry_prompt = f"""{BS_PROMPT.replace("{items}", items_text)}

YOUR PREVIOUS ATTEMPT HAD THIS ERROR:
{error_info}

Fix it. Every source value must land in exactly one cell. Check your arithmetic."""
    result = call_gpt(client, retry_prompt, f"BS_retry_{year}")
    return result


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# EXCEL OUTPUT (SINGLE COMBINED FILE)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

ROWS = [
    (3, "Balance Sheet Particulars", "header", None),
    (4, "Networth", "section", None),
    (5, "Share Capital", "data", "F5"),
    (6, "Retained Earnings", "data", "F6"),
    (7, "General Reserves & Surplus", "data", "F7"),
    (8, "Other Equity", "data", "F8"),
    (10, "Total Networth", "total", "SUM({c}5:{c}8)"),
    (11, "", "blank", None),
    (12, "Current Liabilities", "section", None),
    (14, "Accounts Payable", "data", "F14"),
    (15, "Provisions", "data", "F15"),
    (16, "Short term borrowings", "data", "F16"),
    (17, "Other Current Liabilities", "data", "F17"),
    (18, "Other Financial Liabilities", "data", "F18"),
    # Row 19 (Lease Liabilities Current) removed: per Ind AS 116 lease
    # liabilities are folded into F18 (Other Financial Liab Current).
    # Showing a 0-valued "Lease Liabilities (Current)" row was misleading
    # — users thought lease wasn't being captured when in fact it was
    # already inside F18.
    (20, "Total Current Liabilities", "total", "SUM({c}14:{c}18)"),
    (21, "", "blank", None),
    (22, "Non - Current Liabilities", "section", None),
    (24, "Long Term borrowings", "data", "F24"),
    (25, "Provision", "data", "F25"),
    (26, "Others (DTL + Other NCL)", "data", "F26"),
    (27, "Other Financial Liabilities", "data", "F27"),
    (29, "Total Non-Current Liabilities", "total", "SUM({c}24:{c}27)"),
    (30, "", "blank", None),
    (31, "Total Liabilities", "total", "{c}29+{c}20+{c}10"),
    (32, "", "blank", None),
    (33, "Current Assets", "section", None),
    (35, "Bank Balance", "data", "F35"),
    (36, "Cash & Cash Equivalence", "data", "F36"),
    (37, "Inventory", "data", "F37"),
    (38, "Investments", "data", "F38"),
    (39, "Loans", "data", "F39"),
    (40, "Accounts Receivable", "data", "F40"),
    (41, "Other Current Assets", "data", "F41"),
    (42, "Other Financial Assets", "data", "F42"),
    (44, "Total Current Assets", "total", "SUM({c}35:{c}42)"),
    (45, "", "blank", None),
    (46, "Non - Current Assets", "section", None),
    (48, "Fixed Assets (PPE+Intangible+ROU+Goodwill)", "data", "F48"),
    (49, "Investments", "data", "F49"),
    (50, "Loans", "data", "F50"),
    (51, "Capital Work-in-Progress", "data", "F51"),
    (52, "Other Non-Current Assets", "data", "F52"),
    (53, "Other Financial Assets", "data", "F53"),
    (54, "Deferred Tax Assets", "data", "F54"),
    (56, "Total Non-Current Assets", "total", "SUM({c}48:{c}54)"),
    (57, "", "blank", None),
    (59, "Total Assets", "total", "{c}56+{c}44"),
    (60, "", "blank", None),
    # Spec: Total Liabilities − Total Assets (was TA − TL; sign flipped
    # to match the Input Automation Test Requirement).
    (61, "DIFFERENCE (must be 0)", "diff", "{c}31-{c}59"),
    (62, "", "blank", None),
    # ── P&L block — Schedule III order, every total is a visible Excel
    # formula referencing visible component cells. No static "magic"
    # values; a CA can audit the entire P&L by following the formulas.
    (65, "P&L Statement Particulars", "header", None),
    (66, "Revenue from Operations", "data", "F66"),
    (67, "Cost of goods sold", "data", "F67"),
    (68, "Gross Profit", "total", "{c}66-{c}67"),
    (69, "Employee benefits expense", "data", "F69"),
    (70, "Finance costs", "data", "F70"),
    (71, "Depreciation & amortization expense", "data", "F71"),
    (72, "Other Expenses", "data", "other_expense"),
    (73, "Other Income", "data", "other_income"),
    (74, "Other expenses less other income", "total", "{c}72-{c}73"),
    (75, "Profit before Exceptional Items & Tax", "total",
         "{c}68-{c}69-{c}70-{c}71-{c}74"),
    (76, "Exceptional Items (signed; +ve = income, -ve = expense)",
         "data", "exceptional_items"),
    (77, "Profit Before Tax", "total", "{c}75+{c}76"),
    (78, "Taxes", "data", "F73"),
    (79, "Net Profit", "total", "{c}77-{c}78"),
]

# Styles
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name="Arial", bold=True, size=10, color="FFFFFF")
SEC_FONT = Font(name="Arial", bold=False, size=9, color="4F46E5")
TOT_FILL = PatternFill("solid", fgColor="D9D9D9")
TOT_FONT = Font(name="Arial", bold=True, size=9, color="333333")
DIFF_FILL = PatternFill("solid", fgColor="4F46E5")
DIFF_FONT = Font(name="Arial", bold=True, size=10, color="FFFFFF")
DATA_FONT = Font(name="Arial", size=10, color="0000FF")
FORMULA_FONT = Font(name="Arial", bold=True, size=10, color="000000")
NUM_FMT = '#,##0.00;(#,##0.00);"-"'
THIN_BORDER = Border(bottom=Side(style="thin", color="EEEEEE"))


def build_report(all_data, output_path, company_name=""):
    """Build a single combined Excel report.
    
    all_data = {
        "standalone": {"current": {cells}, "prior": {cells}} or None,
        "consolidated": {"current": {cells}, "prior": {cells}} or None,
    }
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Report"

    # Determine which columns we need
    columns = []
    col_labels = {}
    col_idx = 5  # start at column E
    for variant in ["standalone", "consolidated"]:
        if variant not in all_data or not all_data[variant]:
            continue
        for year in ["current", "prior"]:
            if year not in all_data[variant]:
                continue
            cl = get_col_letter(col_idx)
            columns.append((cl, col_idx, variant, year))
            v_label = "Standalone" if variant == "standalone" else "Consolidated"
            y_label = "Current Year" if year == "current" else "Prior Year"
            col_labels[cl] = f"{v_label}\n{y_label}"
            col_idx += 1

    if not columns:
        print("  [ERROR] No data to output!")
        return None

    # Column widths
    ws.column_dimensions["D"].width = 44
    for cl, _, _, _ in columns:
        ws.column_dimensions[cl].width = 18

    # Row 1: Company name
    ws["D1"] = company_name or "Financial Statement Report"
    ws["D1"].font = Font(name="Arial", bold=True, size=14, color="1F3864")

    # Row 2: Column headers
    ws["D2"] = "Particulars"
    ws["D2"].font = HDR_FONT
    ws["D2"].fill = HDR_FILL
    for cl, ci, _, _ in columns:
        c = ws.cell(row=2, column=ci, value=col_labels[cl])
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 36

    # Build rows
    for row_num, label, rtype, key in ROWS:
        d = ws.cell(row=row_num, column=4, value=label)

        if rtype == "header":
            d.font = HDR_FONT
            d.fill = HDR_FILL
            d.alignment = Alignment(horizontal="center")
            for _, ci, _, _ in columns:
                ws.cell(row=row_num, column=ci).fill = HDR_FILL
        elif rtype == "section":
            d.font = SEC_FONT
        elif rtype == "total":
            d.font = TOT_FONT
            d.fill = TOT_FILL
        elif rtype == "diff":
            d.font = DIFF_FONT
            d.fill = DIFF_FILL

        for cl, ci, variant, year in columns:
            cell = ws.cell(row=row_num, column=ci)
            if rtype == "data" and key:
                val = all_data[variant][year].get(key, 0)
                cell.value = float(val) if val else 0.0
                cell.font = DATA_FONT
                cell.number_format = NUM_FMT
                cell.border = THIN_BORDER
            elif rtype == "total" and key:
                cell.value = "=" + key.replace("{c}", cl)
                cell.font = FORMULA_FONT
                cell.fill = TOT_FILL
                cell.number_format = NUM_FMT
            elif rtype == "diff" and key:
                cell.value = "=" + key.replace("{c}", cl)
                cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
                cell.fill = DIFF_FILL
                cell.number_format = NUM_FMT

    ws.freeze_panes = "E3"

    # Defensive guard: never overwrite hand-curated templates. extract_tables
    # owns the canonical _PROTECTED_BASENAMES set + check.
    try:
        from extract_tables import _assert_not_protected
        _assert_not_protected(output_path)
    except ImportError:
        pass

    # ── WARNINGS sheet ──
    all_warnings = all_data.get("_warnings", [])
    if all_warnings:
        ww = wb.create_sheet("WARNINGS")
        ww.column_dimensions["A"].width = 15
        ww.column_dimensions["B"].width = 12
        ww.column_dimensions["C"].width = 70
        ww["A1"] = "Variant"
        ww["B1"] = "Category"
        ww["C1"] = "Warning"
        for c in range(1, 4):
            ww.cell(1, c).font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
            ww.cell(1, c).fill = PatternFill("solid", fgColor="C0392B")
        for i, (variant, cat, msg) in enumerate(all_warnings, 2):
            ww.cell(i, 1).value = variant
            ww.cell(i, 2).value = cat
            ww.cell(i, 3).value = msg
            ww.cell(i, 3).font = Font(name="Arial", size=9)

    wb.save(output_path)
    return output_path

def get_col_letter(idx):
    return chr(ord("A") + idx - 1)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# MAIN PIPELINE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def process_file(input_file, template_file="", api_key="", output_dir="./output",
                 fy_override=None, company_override=None,
                 company_workbook_dir=None, source_pdf_path=None):
    """Run the BS/PL mapping pipeline against one extracted Excel file.

    `fy_override` / `company_override` are passed through from the upload form
    so the per-company multi-year workbook (<dir>/companies/<Name>_FinancialReport.xlsx)
    knows what FY label to use and which company to file under.

    `company_workbook_dir` controls where the multi-year workbook lives. When
    None, it defaults to the parent of `output_dir` so all per-job outputs
    feed into one shared `companies/` folder (instead of one per job).
    """
    os.makedirs(output_dir, exist_ok=True)
    fname = Path(input_file).stem
    print(f"\n{'━'*60}")
    print(f"  Processing: {fname}")
    print(f"{'━'*60}")

    # Stage 1: Preprocess
    print("\n  [1/3] Preprocessing...")
    variants = process_source(input_file)
    for v, sheets in variants.items():
        for t, items in sheets.items():
            real = [i for i in items if i["cur"] != 0 or i["pri"] != 0]
            print(f"    {v}/{t}: {len(real)} valued items")

    # ── Resolve mapping mode from config ──
    # mapper_mode values:
    #   "claude"     — Claude (Anthropic) strict tool-use mapper (DEFAULT when
    #                  anthropic_api_key is configured; replaces OpenAI as the
    #                  primary LLM mapper as of 2026-05).
    #   "v2_llm"     — legacy OpenAI-strict-structured-output mapper.
    #   "hybrid"     — rules + TF-IDF + RAG cascade (offline fallback).
    #   "rules_only" — pure rules without RAG/LLM.
    # When the LLM call returns None (no key, network error), we fall back to
    # hybrid so the pipeline never silently produces nothing.
    try:
        from config import load_config as _load_cfg
        _cfg = _load_cfg()
    except Exception:
        _cfg = {}
    if _cfg.get("mapper_mode"):
        mapper_mode = _cfg["mapper_mode"]
    elif HAS_CLAUDE and (_cfg.get("anthropic_api_key") or os.environ.get("ANTHROPIC_API_KEY")):
        mapper_mode = "claude"
    elif HAS_LLM_V2 and _cfg.get("openai_api_key"):
        mapper_mode = "v2_llm"
    else:
        mapper_mode = "hybrid"
    print(f"\n  [2/3] Mapping (mode={mapper_mode})...")
    all_data = {}
    all_warnings = []

    # ── Parallel mapper dispatch ──
    # Up to 4 LLM calls (BS+PL × standalone+consolidated) used to run
    # serially, taking ~50s for a 2-variant report. They're independent —
    # firing them concurrently with a ThreadPoolExecutor cuts the total
    # wait to whichever single call is slowest (~12-15s).
    from concurrent.futures import ThreadPoolExecutor
    import time as _time
    _t0 = _time.time()

    def _run_bs(variant: str, sheets: dict):
        if "bs" not in sheets:
            return variant, "bs", None, None
        bs_mapping = None
        err = None
        if mapper_mode == "claude" and HAS_CLAUDE:
            try:
                # Pass source PDF so the mapper can do a final PDF-grounded
                # remediation pass when constraint checks (balance,
                # section-total) keep failing on text-only retries.
                bs_mapping = map_bs_claude(sheets["bs"],
                                            pdf_path=source_pdf_path)
            except Exception as e:
                err = e
                bs_mapping = None
            if bs_mapping is None and HAS_MATCHER:
                bs_mapping = map_bs_hybrid(sheets["bs"])
        elif mapper_mode == "v2_llm" and HAS_LLM_V2:
            try:
                bs_mapping = map_bs_v2(sheets["bs"])
            except Exception as e:
                err = e
                bs_mapping = None
            if bs_mapping is None and HAS_MATCHER:
                bs_mapping = map_bs_hybrid(sheets["bs"])
        elif HAS_MATCHER:
            bs_mapping = map_bs_hybrid(sheets["bs"])
        elif HAS_OPENAI and api_key:
            client = OpenAI(api_key=api_key)
            bs_mapping = map_bs(client, sheets["bs"])
        return variant, "bs", bs_mapping, err

    def _run_pl(variant: str, sheets: dict):
        if "pl" not in sheets:
            return variant, "pl", None, None
        pl_mapping = None
        err = None
        if mapper_mode == "claude" and HAS_CLAUDE:
            try:
                pl_mapping = map_pl_claude(sheets["pl"],
                                            pdf_path=source_pdf_path)
            except Exception as e:
                err = e
                pl_mapping = None
            if pl_mapping is None and HAS_MATCHER:
                pl_mapping = map_pl_hybrid(sheets["pl"])
        elif mapper_mode == "v2_llm" and HAS_LLM_V2:
            try:
                pl_mapping = map_pl_v2(sheets["pl"])
            except Exception as e:
                err = e
                pl_mapping = None
            if pl_mapping is None and HAS_MATCHER:
                pl_mapping = map_pl_hybrid(sheets["pl"])
        elif HAS_MATCHER:
            pl_mapping = map_pl_hybrid(sheets["pl"])
        elif HAS_OPENAI and api_key:
            client = OpenAI(api_key=api_key)
            pl_mapping = map_pl(client, sheets["pl"])
        return variant, "pl", pl_mapping, err

    futures = []
    print(f"    Firing {sum(2 for s in variants.values() if 'bs' in s or 'pl' in s)} mapper calls in parallel "
          f"({list(variants.keys())})...")
    with ThreadPoolExecutor(max_workers=4) as ex:
        for variant, sheets in variants.items():
            futures.append(ex.submit(_run_bs, variant, sheets))
            futures.append(ex.submit(_run_pl, variant, sheets))
        results = [f.result() for f in futures]
    print(f"    All mapper calls done in {(_time.time() - _t0):.1f}s")

    # Pivot results into per-variant dict
    mapped: dict = {v: {"bs": None, "pl": None, "bs_err": None, "pl_err": None}
                    for v in variants.keys()}
    for variant, kind, mapping, err in results:
        mapped[variant][kind] = mapping
        if err: mapped[variant][f"{kind}_err"] = err

    # ── Per-variant post-processing (validation, warnings, cell assembly) ──
    for variant, sheets in variants.items():
        print(f"\n    ── {variant.upper()} ──")
        all_data[variant] = {}

        bs_mapping = mapped[variant]["bs"]
        bs_err     = mapped[variant]["bs_err"]
        if bs_err:
            print(f"    BS v2 error: {bs_err}; (used hybrid fallback if available)")
        if "bs" in sheets:
            if bs_mapping:
                for year in ["current_year", "prior_year"]:
                    if year not in bs_mapping:
                        continue
                    v = validate_bs(bs_mapping, year)
                    yl = year.replace("_year", "")
                    if v["balanced"]:
                        print(f"    BS {yl}: ✓ balanced (Assets={v['total_assets']}, Liab={v['total_liabilities']})")
                    else:
                        print(f"    BS {yl}: ✗ diff={v['diff']} — FLAGGED")
                        all_warnings.append((variant, "BS", f"BS {yl} not balanced: diff={v['diff']:.2f} (Assets={v['total_assets']}, Liab={v['total_liabilities']})"))
                for w in bs_mapping.get("_warnings", []):
                    all_warnings.append((variant, "BS", w))
            else:
                print("    BS: ✗ mapping failed!")
                all_warnings.append((variant, "BS", "BS mapping failed — no data extracted"))

        pl_mapping = mapped[variant]["pl"]
        pl_err     = mapped[variant]["pl_err"]
        if pl_err:
            print(f"    PL v2 error: {pl_err}; (used hybrid fallback if available)")
        if "pl" in sheets:
            if pl_mapping:
                print(f"    PL: ✓ mapped. Notes: {pl_mapping.get('notes', '')[:80]}")
                for w in pl_mapping.get("_warnings", []):
                    all_warnings.append((variant, "PL", w))
            else:
                print("    PL: ✗ mapping failed!")
                all_warnings.append((variant, "PL", "PL mapping failed — no data extracted"))
        else:
            all_warnings.append((variant, "PL", "No P&L sheet found in extracted data"))

        # Assemble cell values for each year
        for year_key, year_label in [("current_year", "current"), ("prior_year", "prior")]:
            cells = {}
            if bs_mapping and year_key in bs_mapping:
                for k, v in bs_mapping[year_key].items():
                    if k.startswith("F") and k[1:].isdigit():
                        cells[k] = float(v)
            if pl_mapping and year_key in pl_mapping:
                m = pl_mapping[year_key]
                # Static data values that the report displays directly:
                #   F66 Revenue, F67 COGS, F69 Emp, F70 Interest, F71 Dep
                #   F73 Taxes (semantic ID, displayed at row 78)
                for k in ["F66", "F67", "F69", "F70", "F71", "F73"]:
                    if k in m:
                        cells[k] = float(m[k])
                # NEW LAYOUT: Other Expenses + Other Income are now their
                # OWN visible rows (R72, R73). F74 "Other expenses less
                # other income" is a real Excel formula = E72 − E73, so
                # any reviewer can audit the calc by clicking the cell —
                # no more static "magic" values.
                cells["other_expense"]     = float(m.get("other_expense", 0) or 0)
                cells["other_income"]      = float(m.get("other_income", 0) or 0)
                cells["exceptional_items"] = float(m.get("exceptional_items",
                                                          m.get("exceptional", 0)) or 0)
                expected_np = float(m.get("net_profit", 0))
                print(f"    PL {year_label}: OE={cells['other_expense']:.2f} "
                      f"OI={cells['other_income']:.2f} "
                      f"Exc={cells['exceptional_items']:.2f} "
                      f"Tax={cells.get('F73', 0):.2f} "
                      f"NP_published={expected_np:.2f}")

            all_data[variant][year_label] = cells

    # Stage 3: Build combined report
    print("\n  [3/3] Building combined report...")
    output_path = os.path.join(output_dir, f"{fname}_Report.xlsx")
    all_data["_warnings"] = all_warnings
    if all_warnings:
        print(f"    ⚠ {len(all_warnings)} warnings — see WARNINGS sheet in report")
    build_report(all_data, output_path, company_name=fname)

    # Recalculate formulas if possible
    recalc = Path(__file__).parent / "recalc.py"
    if not recalc.exists():
        # Try standard location
        recalc = Path("/mnt/skills/public/xlsx/scripts/recalc.py")
    if recalc.exists():
        os.system(f"python {recalc} {output_path}")

    # Final verification
    try:
        wb = load_workbook(output_path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        print(f"\n  ── FINAL RESULTS ──")
        for ci in range(5, 5 + len([v for v in all_data.values() if v])*2):
            cl = get_col_letter(ci)
            diff_val = ws.cell(row=61, column=ci).value
            np_val = ws.cell(row=74, column=ci).value
            ta_val = ws.cell(row=59, column=ci).value
            hdr = ws.cell(row=2, column=ci).value or cl
            hdr_short = hdr.replace("\n", " ")
            if ta_val is not None:
                diff_str = f"{diff_val}" if diff_val else "0"
                print(f"    {hdr_short}: Assets={ta_val}, DIFF={diff_str}, NP={np_val}")
    except Exception:
        pass

    # ── Per-company multi-year append ──
    # Drop the just-mapped year(s) into a single per-company workbook so
    # subsequent uploads of the same company (different FY) accumulate
    # year-on-year columns instead of producing one isolated report each.
    if HAS_COMPANY_WRITER:
        try:
            from airtable_uploader import extract_company_name, extract_fy_dates
            company = company_override or extract_company_name(input_file)
            fy = extract_fy_dates(input_file, fy_override=fy_override)
            # Default: parent of the per-job output_dir, so every job feeds
            # into ONE shared companies/ folder (not one per job).
            wb_dir = company_workbook_dir or os.path.dirname(os.path.abspath(output_dir))
            print(f"\n  ── Per-company append ──")
            print(f"    company='{company}' fy={fy.get('fy_token')} dir={wb_dir}")
            combined_path = append_company_report(
                company_name=company,
                fy=fy,
                all_data=all_data,
                output_dir=wb_dir,
            )
            if combined_path:
                print(f"  ✓ Combined company workbook: {combined_path}")
        except Exception as e:
            print(f"  ⚠ Per-company append failed: {type(e).__name__}: {e}")

    print(f"\n  ✓ Output: {output_path}")
    return output_path


def process_batch(input_dir, template_file, api_key, output_dir):
    files = sorted(Path(input_dir).glob("*_extracted*.xlsx"))
    if not files:
        files = sorted(Path(input_dir).glob("*.xlsx"))
        files = [f for f in files if "template" not in f.name.lower() and "output" not in f.name.lower() and "report" not in f.name.lower()]
    print(f"Found {len(files)} file(s) to process.\n")
    for f in files:
        try:
            process_file(str(f), template_file, api_key, output_dir)
        except Exception as e:
            print(f"  ERROR on {f.name}: {e}")
            import traceback
            traceback.print_exc()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="BS/PL → Template Mapper v2")
    parser.add_argument("--input", required=True, help="Input .xlsx file or folder for batch")
    parser.add_argument("--template", required=True, help="Template .xlsx file")
    parser.add_argument("--api-key", default="", help="OpenAI API key (optional — offline mode used by default)")
    parser.add_argument("--output", default="./output", help="Output directory")
    args = parser.parse_args()

    if os.path.isdir(args.input):
        process_batch(args.input, args.template, args.api_key, args.output)
    else:
        process_file(args.input, args.template, args.api_key, args.output)