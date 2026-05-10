"""
company_excel_writer.py — Per-company multi-year financial template that
APPENDS year columns instead of overwriting on every upload.

Why this exists:
  The original `bs_pl_mapper.build_report` writes one fresh `*_Report.xlsx`
  per pipeline job, with columns for the SINGLE year that was uploaded
  (Standalone CY/PY + optional Consolidated CY/PY). Every subsequent upload
  for the same company creates another standalone file — there's no
  consolidated history view.

  This module solves that. For every successful job we drop into a single
  per-company workbook at:

      <output_dir>/companies/<Company_Name>_FinancialReport.xlsx

  When the same company is uploaded again with a different FY, the new
  year(s) are added as additional columns. If a year already exists in the
  workbook, those columns are UPDATED (not duplicated).

Layout:
  Column D = "Particulars" (the standardized template labels — same row
             order as bs_pl_mapper.ROWS).
  Columns E onwards = one column per (variant, fiscal_year) pair, sorted
             chronologically by end_date so newer years sit further right.
  Header row (row 2) shows e.g. "Standalone\nFY2425" or "Consolidated\nFY2526".
  Row 1 holds the company name.

Public API:
  append_company_report(company_name, fy, all_data, output_dir)
    -> returns the absolute path to the per-company workbook.

  fy is the dict produced by airtable_uploader.extract_fy_dates — we use
  fy["fy_token"] (e.g. "FY2425") as the year label and fy["end_date_current"]
  / "end_date_prior" so the workbook can sort columns by date.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Template rows — same shape as bs_pl_mapper.ROWS (kept here so this module
# is self-contained). If bs_pl_mapper.ROWS evolves, mirror it here.
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

ROWS: List[Tuple[int, str, str, Optional[str]]] = [
    (3,  "Balance Sheet Particulars", "header", None),
    (4,  "Networth", "section", None),
    (5,  "Share Capital", "data", "F5"),
    (6,  "Retained Earnings", "data", "F6"),
    (7,  "General Reserves & Surplus", "data", "F7"),
    (8,  "Other Equity", "data", "F8"),
    (10, "Total Networth", "total", "SUM({c}5:{c}8)"),
    (11, "", "blank", None),
    (12, "Current Liabilities", "section", None),
    (14, "Accounts Payable", "data", "F14"),
    (15, "Provisions", "data", "F15"),
    (16, "Short term borrowings", "data", "F16"),
    (17, "Other Current Liabilities", "data", "F17"),
    (18, "Other Financial Liabilities", "data", "F18"),
    (20, "Total Current Liabilities", "total", "SUM({c}14:{c}18)"),
    (21, "", "blank", None),
    (22, "Non - Current Liabilities", "section", None),
    (24, "Long Term borrowings", "data", "F24"),
    (25, "Provision", "data", "F25"),
    (26, "Others (DTL + Other NCL)", "data", "F26"),
    (27, "Other Financial Liabilities", "data", "F27"),
    (29, "Total Non-Current Liabilities", "total", "SUM({c}24:{c}27)"),
    (30, "", "blank", None),
    (31, "Total Liabilities", "total", "{c}29+{c}20+{c}10"),
    (32, "", "blank", None),
    (33, "Current Assets", "section", None),
    (35, "Bank Balance", "data", "F35"),
    (36, "Cash & Cash Equivalence", "data", "F36"),
    (37, "Inventory", "data", "F37"),
    (38, "Investments", "data", "F38"),
    (39, "Loans", "data", "F39"),
    (40, "Accounts Receivable", "data", "F40"),
    (41, "Other Current Assets", "data", "F41"),
    (42, "Other Financial Assets", "data", "F42"),
    (44, "Total Current Assets", "total", "SUM({c}35:{c}42)"),
    (45, "", "blank", None),
    (46, "Non - Current Assets", "section", None),
    (48, "Fixed Assets (PPE+Intangible+ROU+Goodwill)", "data", "F48"),
    (49, "Investments", "data", "F49"),
    (50, "Loans", "data", "F50"),
    (51, "Capital Work-in-Progress", "data", "F51"),
    (52, "Other Non-Current Assets", "data", "F52"),
    (53, "Other Financial Assets", "data", "F53"),
    (54, "Deferred Tax Assets", "data", "F54"),
    (56, "Total Non-Current Assets", "total", "SUM({c}48:{c}54)"),
    (57, "", "blank", None),
    (59, "Total Assets", "total", "{c}56+{c}44"),
    (60, "", "blank", None),
    (61, "DIFFERENCE (must be 0)", "diff", "{c}31-{c}59"),
    (62, "", "blank", None),
    # ── P&L block — Schedule III order, mirrors bs_pl_mapper.ROWS so
    # the per-job Report and per-company workbook share one canonical
    # layout. Every total is a visible Excel formula.
    (65, "P&L Statement Particulars", "header", None),
    (66, "Revenue from Operations", "data", "F66"),
    (67, "Cost of goods sold", "data", "F67"),
    (68, "Gross Profit", "total", "{c}66-{c}67"),
    (69, "Employee benefits expense", "data", "F69"),
    (70, "Finance costs", "data", "F70"),
    (71, "Depreciation & amortization expense", "data", "F71"),
    (72, "Other Expenses", "data", "other_expense"),
    (73, "Other Income", "data", "other_income"),
    (74, "Other expenses less other income", "total", "{c}72-{c}73"),
    (75, "Profit before Exceptional Items & Tax", "total",
         "{c}68-{c}69-{c}70-{c}71-{c}74"),
    (76, "Exceptional Items (signed; +ve = income, -ve = expense)",
         "data", "exceptional_items"),
    (77, "Profit Before Tax", "total", "{c}75+{c}76"),
    (78, "Taxes", "data", "F73"),
    (79, "Net Profit", "total", "{c}77-{c}78"),
]


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Styles
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

HDR_FILL    = PatternFill("solid", fgColor="1F3864")
HDR_FONT    = Font(name="Arial", bold=True, size=10, color="FFFFFF")
SEC_FONT    = Font(name="Arial", bold=False, size=9, color="4F46E5")
TOT_FILL    = PatternFill("solid", fgColor="D9D9D9")
TOT_FONT    = Font(name="Arial", bold=True, size=9, color="333333")
DIFF_FILL   = PatternFill("solid", fgColor="4F46E5")
DIFF_FONT   = Font(name="Arial", bold=True, size=10, color="FFFFFF")
DATA_FONT   = Font(name="Arial", size=10, color="0000FF")
FORMULA_FONT = Font(name="Arial", bold=True, size=10, color="000000")
NUM_FMT     = '#,##0.00;(#,##0.00);"-"'
THIN_BORDER = Border(bottom=Side(style="thin", color="EEEEEE"))


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Helpers
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

_COMPANY_NAME_RE = re.compile(r"[^\w\s\-]+", re.UNICODE)


def _safe_filename(name: str) -> str:
    name = _COMPANY_NAME_RE.sub("", name).strip()
    name = re.sub(r"\s+", "_", name)
    return name or "Company"


def _fy_year_from_token(fy_token: str) -> int:
    """FY2425 → 2025 (the END year). Used as a chronological sort key."""
    m = re.match(r"FY(\d{2})(\d{2})", fy_token or "")
    if not m:
        return 0
    return 2000 + int(m.group(2))


def _fy_human_label(fy_token: str) -> str:
    """FY2425 → '2024-25'. Used for the column header so end users don't
    have to mentally decode 'FY2425'. Keeps both formats for clarity."""
    m = re.match(r"FY(\d{2})(\d{2})", fy_token or "")
    if not m:
        return fy_token or ""
    return f"20{m.group(1)}-{m.group(2)}"


def _column_label(variant: str, fy_token: str) -> str:
    """Two-line header: 'Standalone\\nFY 2024-25'. The Variant goes on
    line 1 (Standalone vs Consolidated), the human-readable financial
    year on line 2 so it's obvious which year each column represents.

    Header text is also the dedup key — same string ⇒ same column gets
    overwritten on re-upload, different string ⇒ new column to the right."""
    return f"{variant.capitalize()}\nFY {_fy_human_label(fy_token)}"


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Workbook reader — reconstruct existing year-columns from an existing file
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _data_row_for_cell(cell_key: str) -> Optional[int]:
    """F35 → 35. Used to walk the template column."""
    m = re.match(r"^F(\d+)$", cell_key or "")
    return int(m.group(1)) if m else None


_OLD_HEADER_RE = re.compile(r"^(Standalone|Consolidated)\s*[\r\n]+\s*FY(\d{2})(\d{2})\s*$",
                             re.IGNORECASE)


def _normalize_header(raw: str) -> str:
    """Migrate the legacy 'Standalone\\nFY2425' header to the new
    'Standalone\\nFY 2024-25' format so existing per-company workbooks
    don't sprout duplicate columns when re-saved with the new code."""
    s = (raw or "").strip()
    m = _OLD_HEADER_RE.match(s)
    if m:
        variant = m.group(1).capitalize()
        token = f"FY{m.group(2)}{m.group(3)}"
        return _column_label(variant.lower(), token)
    return s


def _read_existing_columns(ws) -> Dict[str, Dict[str, float]]:
    """Walk an already-built per-company workbook and recover the data
    columns as {column_label: {F-cell: value}}.

    Header row 2 has multi-line strings like 'Standalone\\nFY 2024-25'.
    Old workbooks (pre-2026-05) used 'Standalone\\nFY2425' — those headers
    are normalized on the fly so a re-save migrates without duplication.
    """
    out: Dict[str, Dict[str, float]] = {}
    if ws.max_row < 2:
        return out
    for col_idx in range(5, ws.max_column + 1):
        header = ws.cell(row=2, column=col_idx).value
        if not header:
            continue
        key = _normalize_header(str(header))
        if not key:
            continue
        col_data: Dict[str, float] = {}
        for _row, _label, rtype, cell_key in ROWS:
            if rtype != "data" or not cell_key:
                continue
            data_row = _data_row_for_cell(cell_key)
            if data_row is None:
                continue
            v = ws.cell(row=data_row, column=col_idx).value
            if v is None:
                continue
            try:
                col_data[cell_key] = float(v)
            except (TypeError, ValueError):
                # Skip formula strings — those are recomputed on write.
                pass
        out[key] = col_data
    return out


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Public API
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def append_company_report(company_name: str,
                           fy: Dict,
                           all_data: Dict,
                           output_dir: str | Path) -> Optional[str]:
    """Append the new year(s) of data to the per-company combined workbook.

    Args:
      company_name : display name (e.g. "Akiko Global Services").
      fy           : dict from airtable_uploader.extract_fy_dates — must have
                     keys 'fy_token', 'end_date_current', 'end_date_prior'.
                     Both PY and CY columns are added; the PY end-year sorts
                     before the CY end-year so columns end up chronological.
      all_data     : the same dict that bs_pl_mapper.build_report consumes
                     ({"standalone": {"current": {...}, "prior": {...}},
                       "consolidated": ..., "_warnings": [...]}).
      output_dir   : root output directory. Workbook is written under
                     <output_dir>/companies/<safe_name>_FinancialReport.xlsx.

    Returns the absolute path to the workbook, or None if there was nothing
    to append.
    """
    if not company_name:
        return None
    if not fy or not fy.get("fy_token"):
        return None

    # Where does each (variant, year) live in `all_data`?
    fy_token = fy["fy_token"]                       # "FY2425"
    cy_end_year = _fy_year_from_token(fy_token)
    py_end_year = cy_end_year - 1
    py_token = f"FY{py_end_year - 2000 - 1:02d}{py_end_year - 2000:02d}"

    # Build the (column_key, sort_key, cells_dict) triples for what this
    # upload contributes. Empty dicts (no data) are skipped so we don't blow
    # away an existing year with zeros.
    incoming: List[Tuple[str, int, Dict[str, float]]] = []
    for variant in ("standalone", "consolidated"):
        bucket = all_data.get(variant) or {}
        for year_kind, token, sort_year in (
            ("current", fy_token, cy_end_year),
            ("prior",   py_token, py_end_year),
        ):
            cells = bucket.get(year_kind) or {}
            if not any(float(v or 0) for v in cells.values()):
                continue
            label = _column_label(variant, token)
            # Sort within the same FY: standalone first, then consolidated.
            sort_tuple = sort_year * 10 + (0 if variant == "standalone" else 1)
            incoming.append((label, sort_tuple, dict(cells)))

    if not incoming:
        print(f"  [company_excel] nothing to append for {company_name}")
        return None

    out_root = Path(output_dir) / "companies"
    out_root.mkdir(parents=True, exist_ok=True)
    file_path = out_root / f"{_safe_filename(company_name)}_FinancialReport.xlsx"

    # ── Merge incoming with what's already in the file ──
    columns: Dict[str, Tuple[int, Dict[str, float]]] = {}
    if file_path.exists():
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            existing = _read_existing_columns(ws)
            for k, cells in existing.items():
                # Pull the sort key from the FY in the header label. Match
                # both the new "FY 2024-25" form and the legacy "FY2425".
                m = re.search(r"FY\s*(\d{2,4})\s*-?\s*(\d{2,4})", k)
                end_year = 0
                if m:
                    g2 = m.group(2)
                    end_year = int(g2) if len(g2) == 4 else (2000 + int(g2))
                base = end_year * 10 + (1 if k.lower().startswith("consolidated") else 0)
                columns[k] = (base, cells)
            print(f"  [company_excel] loaded existing workbook with "
                  f"{len(existing)} year-column(s)")
        except Exception as e:
            print(f"  [company_excel] WARN existing workbook unreadable "
                  f"({type(e).__name__}: {e}); rebuilding from scratch")
            columns = {}

    for label, sort_tuple, cells in incoming:
        prev = columns.get(label)
        if prev:
            print(f"  [company_excel] updating existing column '{label.replace(chr(10),' ')}'")
        else:
            print(f"  [company_excel] adding new column '{label.replace(chr(10),' ')}'")
        columns[label] = (sort_tuple, cells)

    # ── Rewrite the workbook from scratch (simpler than surgical edits and
    #    keeps all formulas + styles consistent) ──
    return _write_workbook(file_path, company_name, columns)


def _write_workbook(file_path: Path, company_name: str,
                    columns: Dict[str, Tuple[int, Dict[str, float]]]) -> str:
    # Defensive guard: never overwrite hand-curated templates.
    try:
        from extract_tables import _assert_not_protected
        _assert_not_protected(file_path)
    except ImportError:
        pass

    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Report"

    # Sort columns chronologically (oldest → newest, standalone before
    # consolidated within the same FY).
    sorted_keys = sorted(columns.keys(), key=lambda k: columns[k][0])

    # Layout: D = labels, E.. = year columns
    ws.column_dimensions["D"].width = 44
    for i, _ in enumerate(sorted_keys):
        col = get_column_letter(5 + i)
        ws.column_dimensions[col].width = 18

    ws["D1"] = company_name
    ws["D1"].font = Font(name="Arial", bold=True, size=14, color="1F3864")

    ws["D2"] = "Particulars"
    ws["D2"].font = HDR_FONT
    ws["D2"].fill = HDR_FILL

    for i, key in enumerate(sorted_keys):
        col_idx = 5 + i
        c = ws.cell(row=2, column=col_idx, value=key)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 36

    for row_num, label, rtype, key in ROWS:
        d = ws.cell(row=row_num, column=4, value=label)
        if rtype == "header":
            d.font = HDR_FONT
            d.fill = HDR_FILL
            d.alignment = Alignment(horizontal="center")
            for i in range(len(sorted_keys)):
                ws.cell(row=row_num, column=5 + i).fill = HDR_FILL
        elif rtype == "section":
            d.font = SEC_FONT
        elif rtype == "total":
            d.font = TOT_FONT
            d.fill = TOT_FILL
        elif rtype == "diff":
            d.font = DIFF_FONT
            d.fill = DIFF_FILL

        for i, col_key in enumerate(sorted_keys):
            col_idx = 5 + i
            cl = get_column_letter(col_idx)
            cell = ws.cell(row=row_num, column=col_idx)
            if rtype == "data" and key:
                v = columns[col_key][1].get(key, 0) or 0
                cell.value = float(v)
                cell.font = DATA_FONT
                cell.number_format = NUM_FMT
                cell.border = THIN_BORDER
            elif rtype == "total" and key:
                cell.value = "=" + key.replace("{c}", cl)
                cell.font = FORMULA_FONT
                cell.fill = TOT_FILL
                cell.number_format = NUM_FMT
            elif rtype == "diff" and key:
                cell.value = "=" + key.replace("{c}", cl)
                cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
                cell.fill = DIFF_FILL
                cell.number_format = NUM_FMT

    ws.freeze_panes = "E3"
    wb.save(file_path)
    print(f"  [company_excel] saved {file_path} ({len(sorted_keys)} year-column(s))")
    return str(file_path)
